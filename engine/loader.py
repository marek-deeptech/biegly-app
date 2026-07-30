"""Wczytywanie danych UTP z pliku xlsx do struktur w pamięci.

Czytamy arkusze przez nagłówki kolumn (po nazwie), nie po stałych indeksach —
dzięki temu wariacje układu kolumn nie psują silnika. Każdy wiersz zwracamy
jako dict {nazwa_kolumny: wartość}.
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl


def _match_sheet(wb, name: str):
    """Dopasowanie arkusza po nazwie z tolerancją.

    1) dokładna nazwa (po strip/lower),
    2) prefiks pierwszego członu nazwy docelowej — żeby obsłużyć warianty układu:
       'Transakcje' → 'Transakcje all', 'Zlecenia BO' → 'Zlecenia'.
    """
    target = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == target:
            return wb[sheet_name]
    key = target.split()[0]  # 'transakcje' / 'zlecenia'
    cands = [sn for sn in wb.sheetnames if sn.strip().lower().startswith(key)]
    if cands:
        cands.sort(key=len)  # najkrótsza nazwa = zwykle arkusz główny
        return wb[cands[0]]
    raise KeyError(f"Brak arkusza pasującego do {name!r}; dostępne: {wb.sheetnames}")


def load_rows(path: Path, sheet_name: str) -> list[dict]:
    """Zwraca listę wierszy arkusza jako dicty kluczowane nazwą kolumny.

    Nagłówek = pierwszy wiersz z co najmniej 3 niepustymi komórkami (część
    eksportów GPW ma wiodące puste wiersze nad nagłówkiem, np. arkusz 'Zlecenia').
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _match_sheet(wb, sheet_name)
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = None
        for r in rows:
            cells = [str(h).strip() if h is not None else "" for h in r]
            if sum(1 for c in cells if c) >= 3:
                header = cells
                break
        if header is None:
            return []
        out = [dict(zip(header, r)) for r in rows]
        # Forward-fill DATA_SESJI: część eksportów (np. MLM UTP) wpisuje datę
        # tylko w pierwszym wierszu bloku sesji (konwencja scalonych komórek) —
        # puste wiersze należą do tej samej sesji co ostatnia wpisana data.
        # Pliki w pełni datowane (HubTech) pozostają bez zmiany.
        if "DATA_SESJI" in header:
            last = None
            for row in out:
                v = row.get("DATA_SESJI")
                if v is None or (isinstance(v, str) and not v.strip()):
                    row["DATA_SESJI"] = last
                else:
                    last = v
        return out
    finally:
        wb.close()


def load_trem_paired(source) -> list[dict]:
    """Wczytuje SPAROWANE transakcje TREM (kupno i sprzedaż w jednym wierszu) — akceptuje
    dwa warianty formatu UKNF spotykane w aktach:

      • arkusz ``IAD_C_TREM`` — HubTech/MLM (strona kupna = ``ACCTOWNR_POPRAWIONY_B``),
      • arkusz ``2_stronnie`` — ZASTAL, plik per instrument (``UTP TREM CSY/RSY.xlsx``);
        strona kupna oznaczona sufiksem ``_K`` (Kupno). Plik ma też arkusz ``1_stronnie``
        (jednostronny), którego NIE używamy.

    Stronę kupna normalizujemy do ``ACCTOWNR_POPRAWIONY_B``, którego oczekuje
    ``engine.compute_trem`` — dzięki temu silnik działa dla obu formatów bez zmian.
    ``source`` to bajty albo obiekt plikopodobny. Podnosi ``KeyError``, gdy w pliku nie
    ma żadnego arkusza transakcji sparowanych (np. surowy MiFIR per osoba: ``TREM_Uproszczony``).
    """
    raw = source.read() if hasattr(source, "read") else source
    rows: list[dict] | None = None
    for sheet in ("IAD_C_TREM", "2_stronnie"):
        try:
            candidate = load_rows(io.BytesIO(raw), sheet)  # świeży strumień na każdą próbę
        except KeyError:
            continue
        if candidate:
            rows = candidate
            break
    if rows is None:
        raise KeyError("Brak arkusza transakcji sparowanych ('IAD_C_TREM' ani '2_stronnie').")
    # Alias kupującego: 2_stronnie ma ACCTOWNR_POPRAWIONY_K (Kupno) zamiast _B.
    for row in rows:
        if row.get("ACCTOWNR_POPRAWIONY_B") is None and row.get("ACCTOWNR_POPRAWIONY_K") is not None:
            row["ACCTOWNR_POPRAWIONY_B"] = row["ACCTOWNR_POPRAWIONY_K"]
    return rows


def load_knf_orders(source, isins: list[str] | None = None) -> tuple[list[dict], dict]:
    """Adapter: ZESTAWIENIE ZLECEŃ KNF → format arkusza „Zlecenia BO" (UTP).

    W aktach ZASTAL nie ma pliku UTP z arkuszem zleceń, ale JEST zbiorcze zestawienie
    zleceń giełdowych podmiotów Grupy sporządzone przez KNF (zał. 5 do zawiadomienia):
    jeden wiersz na zlecenie, kolumny ``Rodzaj zlecenia`` (K/S), ``Data złożenia``
    (z czasem), ``ISIN``, ``Wolumen``, ``Limit``, ``Właściciel``, ``Realizacja``,
    ``Nr rachunku``, ``DM``, ``Mod / Anulata``.

    Mapowanie na klucze silnika (engine.spoofing / engine.metrics):
      Rodzaj zlecenia → ``K/S`` · Data złożenia → ``Data`` + ``OrderEntry Time``
      Wolumen → ``Wolumen`` · Realizacja → ``Wolumen zreal.`` · Limit → ``Limit``
      DM → ``Biuro`` · Nr rachunku → ``Konto``

    ZDARZENIA ANULACJI I MODYFIKACJI. Zestawienie zapisuje je jako OSOBNE WIERSZE
    o rodzaju ``Anulata K``/``Anulata S``/``Modyfikacja K``/``Modyfikacja S``, powiązane
    z pierwotnym zleceniem wspólnym ``Nr zlecenia``; czas zdarzenia stoi w kolumnie
    ``Data złożenia`` tego wiersza. Adapter scala je z pierwotnym zleceniem, wypełniając
    ``CancelReplaceTime`` (anulata) i ``OrderModificationDate`` (modyfikacja) — dzięki temu
    rekonstrukcja śróddzienna zna moment wycofania zlecenia, a nie tylko jego złożenie.
    Zdarzenia bez dopasowanego zlecenia bazowego są pomijane (nie tworzą duplikatów).

    Zwraca ``(orders, owner_map)``; owner_map = {(Biuro, Konto) → Właściciel} zbudowana
    wprost z kolumny ``Właściciel`` (w UTP wymagała głosowania po arkuszu transakcji).
    ``isins`` (opcjonalnie) ogranicza wynik do instrumentów sprawy.
    """
    from .identity import norm_acct  # import lokalny — unika cyklu przy starcie modułu

    raw = source.read() if hasattr(source, "read") else source
    rows = load_rows(io.BytesIO(raw), "Arkusz1")
    want = {str(i).strip().upper() for i in (isins or [])}

    def _ts(v) -> str:
        return v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "strftime") else str(v or "")

    # 1) zdarzenia: Nr zlecenia → (czas anulacji, czas modyfikacji)
    cancels: dict[str, str] = {}
    mods: dict[str, str] = {}
    for r in rows:
        kind = str(r.get("Rodzaj zlecenia") or "").strip().lower()
        nr = str(r.get("Nr zlecenia") or "").strip()
        if not nr:
            continue
        if kind.startswith("anulata"):
            cancels[nr] = _ts(r.get("Data złożenia"))
        elif kind.startswith("modyfikacja"):
            mods[nr] = _ts(r.get("Data złożenia"))

    # 2) zlecenia bazowe (K/S) + doklejenie zdarzeń
    out: list[dict] = []
    owner_map: dict[tuple[str, str], str] = {}
    for r in rows:
        isin = str(r.get("ISIN") or "").strip().upper()
        if want and isin not in want:
            continue
        kind = str(r.get("Rodzaj zlecenia") or "").strip().upper()
        if kind not in ("K", "S"):  # wiersze zdarzeń obsłużone wyżej
            continue
        entry_s = _ts(r.get("Data złożenia"))
        nr = str(r.get("Nr zlecenia") or "").strip()
        biuro, konto = r.get("DM"), r.get("Nr rachunku")
        owner = str(r.get("Właściciel") or "").strip()
        if owner:
            owner_map[(norm_acct(biuro), norm_acct(konto))] = owner
        out.append({
            "Data": entry_s[:10],
            "K/S": kind,
            "Biuro": biuro,
            "Konto": konto,
            "Wolumen": r.get("Wolumen"),
            "Wolumen zreal.": r.get("Realizacja"),
            "Limit": r.get("Limit"),
            "OrderEntry Time": entry_s,
            "CancelReplaceTime": cancels.get(nr, ""),
            "OrderModificationDate": mods.get(nr, ""),
            "ISIN": isin,
            "Właściciel": owner,
            "Nr zlecenia": nr,
        })
    return out, owner_map


def session_date(value) -> str:
    """Normalizuje wartość daty sesji do formatu ISO 'YYYY-MM-DD'."""
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value)[:10]
