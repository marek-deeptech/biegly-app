"""Parser tickowego „widoku arkusza zleceń" GPW → linie BestBid/BestAsk per sesja.

Opcjonalne, dodatkowe źródło do modułu Spoofing & Layering. Gdy w aktach pojawi się
plik z tickowym arkuszem/dziennikiem zleceń, ten parser odtwarza rzeczywiste kwotowania
(BestBid/BestAsk) na tej samej siatce czasu co `engine.spoofing.session_series`, więc
renderer może narysować wykres 1:1 ze wzorem. Bez takiego pliku moduł działa jak dotąd
(obszary wolumenu + kurs transakcyjny) — parser zwraca None i następuje fallback.

Rozpoznawanie tolerancyjne (po nazwach kolumn, warianty PL/EN). Dwa tryby:
  A. MIGAWKI: kolumny czas + BestBid + BestAsk (gotowe kwotowania) → wprost.
  B. DZIENNIK ZDARZEŃ: czas + strona + limit + (id zlecenia) + status (Nowe/Modyfikacja/
     Anulowane/Wykonanie) → replay arkusza (silnik dopasowań): best bid = max cena kupna
     w arkuszu, best ask = min cena sprzedaży; kwotowania z definicji się nie krzyżują.

Każdy błąd/nierozpoznany format → None (bezpieczny fallback).
"""
from __future__ import annotations

from collections import defaultdict

import openpyxl

from .loader import session_date
from .spoofing import CLOSE_S, N_POINTS, OPEN_S, _date, _f, _sec, _time

# Warianty nazw kolumn (dopasowanie: nagłówek zawiera którykolwiek fragment, małe litery).
COLS = {
    "time": ["czas", "time", "godz", "timestamp", "hora"],
    "date": ["data", "date", "dzień", "dzien", "sesja", "session"],
    "side": ["strona", "k/s", "b/s", "side", "kupno/sprzeda"],
    "price": ["limit", "cena", "price", "kurs"],
    "vol": ["wolumen", "wol", "ilość", "ilosc", "volume", "qty", "quantity"],
    "status": ["status", "typ zlec", "typ", "akcja", "action", "rodzaj", "operacja", "event"],
    "oid": ["numer zl", "nr zlec", "numer", "order id", "orderid", "id zlec", "id"],
    "bid": ["bestbid", "best bid", "najlepsza oferta kupna", "najlepszy bid", "bid"],
    "ask": ["bestask", "best ask", "najlepsza oferta sprzeda", "najlepszy ask", "ask", "offer", "oferta"],
}


def _find_header(ws, max_scan: int = 12):
    """Zwraca (row_idx, {pole: col_idx}) dla pierwszego wiersza wyglądającego na nagłówek."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if sum(1 for c in cells if c) < 3:
            continue
        mapping: dict[str, int] = {}
        for field, aliases in COLS.items():
            for j, c in enumerate(cells):
                if c and any(a in c for a in aliases):
                    mapping.setdefault(field, j)
                    break
        # nagłówek uznajemy, gdy jest czas oraz albo (bid i ask), albo (strona i cena)
        if "time" in mapping and (("bid" in mapping and "ask" in mapping) or ("side" in mapping and "price" in mapping)):
            return i, mapping
    return None, None


def _grid():
    step = (CLOSE_S - OPEN_S) / (N_POINTS - 1)
    return [OPEN_S + int(step * i) for i in range(N_POINTS)]


def _carry(arr: list):
    last = None
    for i in range(len(arr)):
        if arr[i] is None:
            arr[i] = last
        else:
            last = arr[i]
    return arr


def _side_norm(v) -> str:
    s = str(v).strip().upper()
    if s.startswith("K") or s.startswith("B"):  # Kupno / Buy
        return "K"
    if s.startswith("S"):  # Sprzedaż / Sell
        return "S"
    return ""


def _status_norm(v) -> str:
    s = str(v).strip().lower()
    if any(k in s for k in ["anul", "cancel", "wycof", "usun", "delete"]):
        return "cancel"
    if any(k in s for k in ["wykon", "real", "trade", "exec", "fill", "transak"]):
        return "exec"
    if any(k in s for k in ["modyf", "modif", "zmian", "replace", " na", "na "]):
        return "modify"
    if any(k in s for k in ["now", "new", "dodan", "add", "wprow", "zlożen", "zlozen"]):
        return "new"
    return "new"


def _sample_snapshots(rows: list[tuple[int, float, float]]) -> tuple[list, list]:
    """rows = (sec, bid, ask) posortowane → wartości na siatce (carry-forward)."""
    rows = sorted(rows)
    grid = _grid()
    bid: list = [None] * N_POINTS
    ask: list = [None] * N_POINTS
    j = 0
    lb = la = None
    for i, t in enumerate(grid):
        while j < len(rows) and rows[j][0] <= t:
            lb, la = rows[j][1], rows[j][2]
            j += 1
        bid[i], ask[i] = lb, la
    return _carry(bid), _carry(ask)


def _replay_events(events: list[dict]) -> tuple[list, list]:
    """events = {sec, side, price, vol, oid, status} posortowane → BestBid/BestAsk na siatce."""
    events.sort(key=lambda e: e["sec"])
    grid = _grid()
    book: dict = {}  # oid -> (side, price)
    bid: list = [None] * N_POINTS
    ask: list = [None] * N_POINTS
    j = 0
    for i, t in enumerate(grid):
        while j < len(events) and events[j]["sec"] <= t:
            e = events[j]
            j += 1
            oid = e["oid"]
            if e["status"] in ("cancel", "exec"):
                book.pop(oid, None)
            elif e["price"] > 0 and e["side"]:
                book[oid] = (e["side"], e["price"])  # new/modify → stan bieżący
        buys = [p for (sd, p) in book.values() if sd == "K"]
        sells = [p for (sd, p) in book.values() if sd == "S"]
        bid[i] = max(buys) if buys else None
        ask[i] = min(sells) if sells else None
    return _carry(bid), _carry(ask)


def parse_orderbook(file, want_days: set[str] | None = None) -> dict | None:
    """Zwraca {day: {"bid": [...N], "ask": [...N]}} dla rozpoznanych dni albo None.

    `want_days` (opcjonalnie) ogranicza parsowanie do wskazanych sesji.
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return None
    try:
        out: dict[str, dict] = {}
        for sh in wb.sheetnames:
            ws = wb[sh]
            hdr, m = _find_header(ws)
            if not m:
                continue
            snap = "bid" in m and "ask" in m
            evt = "side" in m and "price" in m and "status" in m and "oid" in m
            if not (snap or evt):
                continue
            by_day_snap: dict[str, list] = defaultdict(list)
            by_day_evt: dict[str, list] = defaultdict(list)
            cur_day = ""
            for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
                get = lambda f: (row[m[f]] if f in m and m[f] < len(row) else None)  # noqa: E731
                d = _date(get("date")) or (session_date(get("date")) if get("date") else "") or cur_day
                if d:
                    cur_day = d
                tsec = _sec(_time(get("time")))
                if tsec is None:
                    continue
                if want_days and d and d not in want_days:
                    continue
                if snap:
                    b, a = _f(get("bid")), _f(get("ask"))
                    if b > 0 or a > 0:
                        by_day_snap[d or cur_day].append((tsec, b, a))
                else:
                    by_day_evt[d or cur_day].append({
                        "sec": tsec, "side": _side_norm(get("side")), "price": _f(get("price")),
                        "vol": _f(get("vol")), "oid": str(get("oid")), "status": _status_norm(get("status")),
                    })
            for d, rows in by_day_snap.items():
                if d and rows:
                    b, a = _sample_snapshots(rows)
                    out[d] = {"bid": b, "ask": a}
            for d, evs in by_day_evt.items():
                if d and evs and d not in out:
                    b, a = _replay_events(evs)
                    out[d] = {"bid": b, "ask": a}
        return out or None
    finally:
        wb.close()
