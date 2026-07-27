"""Korelacja adresów IP z logowań — deterministyczny dowód „zbieżności IP" (Krok 4).

Źródło: pliki logowań `Logins_users_*.xlsx` (arkusz z kolumnami Username, IpAddress,
Date, Time). Wartości są w formacie FIX, np. `2(Username)=fortune`,
`5(IpAddress)=89.250.20.10` — wyłuskujemy część po znaku `=`.

Wynik: pary użytkowników, którzy logowali się z tych samych adresów IP. To surowa
zbieżność (dowód), nie przesądzenie o koordynacji — interpretuje biegły.
"""
from __future__ import annotations

import io
import re
from collections import defaultdict

import openpyxl

# Wartości w formacie FIX: `tag(Nazwa)=wartość`, np. `2(Username)=fortune`.
_FIX = re.compile(r"^\s*\d+\(([^)]+)\)=(.*)$", re.S)


def _val(cell) -> str:
    """Wyłuskuje wartość z komórki FIX `tag(Nazwa)=wartość` albo zwraca surowy string."""
    if cell is None:
        return ""
    s = str(cell).strip()
    return s.split("=", 1)[1].strip() if "=" in s else s


def _fields(row: dict) -> dict[str, str]:
    """Mapuje wiersz na {nazwa_pola (małe litery): wartość}.

    Preferuje nazwę pola ze znacznika FIX zawartego w treści komórki
    (`(Username)`, `(IpAddress)`) — dzięki temu działa niezależnie od etykiety
    nagłówka kolumny (w części plików „User”, w innych „Username”). Gdy komórka
    nie jest w formacie FIX, kluczem jest nazwa z nagłówka.
    """
    out: dict[str, str] = {}
    for header, cell in row.items():
        if cell is None:
            continue
        s = str(cell).strip()
        m = _FIX.match(s)
        if m:
            out[m.group(1).strip().lower()] = m.group(2).strip()
        elif header:
            out[str(header).strip().lower()] = s
    return out


def load_logins(file) -> list[dict]:
    """Czyta pierwszy arkusz pliku logowań jako listę dictów (nagłówek = 1. wiersz z >=3 komórkami)."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(min_row=1, values_only=True)
        header = None
        for r in it:
            cells = [str(h).strip() if h is not None else "" for h in r]
            if sum(1 for c in cells if c) >= 3:
                header = cells
                break
        if not header:
            return []
        return [dict(zip(header, r)) for r in it]
    finally:
        wb.close()


def _read_grid(data: bytes) -> list[list]:
    """Wiersze pierwszego arkusza jako listy komórek. Obsługuje xlsx/xlsm (openpyxl,
    sygnatura ZIP „PK") oraz stary .xls (xlrd, OLE2). xlrd importujemy leniwie — bez
    plików .xls w ogóle nie jest potrzebny."""
    if data[:2] == b"PK":
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            return [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    import xlrd  # tylko dla starego .xls (OLE2)

    book = xlrd.open_workbook(file_contents=data)
    sh = book.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


_IP_COLS = ("ip", "ip_adr", "adres ip", "adres_ip", "ipaddress", "ip address")
_DATE_COLS = ("start", "data rozpoczęcia", "data rozpoczecia", "od", "date", "data", "login")
_IPV4 = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")


def entity_from_filename(fn: str) -> str:
    """Etykieta podmiotu z NAZWY pliku logowań — w eksportach maklerskich (epromak, DM BOŚ,
    Santander) tożsamość nie jest w danych, tylko w nazwie: „AMIDA CAPITAL_19002399_0914_
    logowania.xls" → „AMIDA CAPITAL", „0902_K.Wieczorek_84217301-logowania IP.xlsx" → „K.Wieczorek"."""
    s = fn.rsplit("/", 1)[-1]
    s = re.sub(r"\.(xlsx?|xlsm|txt)$", "", s, flags=re.I)
    s = re.sub(r"[-_]+", " ", s)  # separatory → spacje NAJPIERW, by \b działało przy numerach
    s = re.sub(r"\s*logowania(\s*ip)?\s*$", "", s, flags=re.I)
    s = re.sub(r"santander\w*", " ", s, flags=re.I)
    s = re.sub(r"dm\s*bo\S*", " ", s, flags=re.I)  # DM BOŚ / „DM BO¦" (mojibake)
    s = re.sub(r"\b\d{4,}\b", " ", s)  # numery rachunków (teraz otoczone spacjami)
    s = re.sub(r"^\s*\d+\s+", "", s)  # wiodący kod konta „0902 "
    s = s.replace("³", "ł").replace("£", "Ł")  # częsty mojibake CP1250 w nazwach (W³odzimierz→Włodzimierz)
    s = re.sub(r"\s{2,}", " ", s).strip(" -_.")
    return s or fn.rsplit("/", 1)[-1]


def load_login_events(data: bytes, filename: str = "") -> list[dict]:
    """Normalizuje logowania z pliku do rekordów {username, ipaddress, date} — jednolite
    wejście do ip_correlation. Rozpoznaje: FIX (tag(Username)=…, tożsamość w pliku —
    HubTech/MLM), kolumnowe eksporty maklerskie (epromak .xls: kol. ip_adr/start; DM BOŚ:
    „Adres Ip"/„Data rozpoczęcia"; prosty ip|od|do) oraz .txt z separatorem „|"
    (rachunek|ip|start|stop|kanał). W eksportach maklerskich username bierzemy z NAZWY PLIKU."""
    ent = entity_from_filename(filename)

    # TXT z separatorem „|": rachunek|ip|start|stop|kanał
    if filename.lower().endswith(".txt"):
        out: list[dict] = []
        for line in data.decode("utf-8", "replace").splitlines():
            parts = line.split("|")
            if len(parts) < 3 or not _IPV4.match(parts[1].strip()):
                continue
            out.append({"username": ent, "ipaddress": parts[1].strip(), "date": _iso_date(parts[2])})
        return out

    grid = _read_grid(data)

    # FIX (tożsamość w pliku): nagłówek = 1. wiersz z ≥3 komórkami, username/ip z tagów FIX.
    if any(c is not None and _FIX.match(str(c).strip()) for row in grid[:40] for c in row):
        header, out = None, []
        for row in grid:
            cells = [str(c).strip() if c is not None else "" for c in row]
            if header is None:
                if sum(1 for c in cells if c) >= 3:
                    header = cells
                continue
            f = _fields(dict(zip(header, row)))
            u = f.get("username") or f.get("user") or f.get("login")
            ip = f.get("ipaddress") or f.get("ip") or f.get("ipaddr") or f.get("adres ip")
            if u and ip:
                out.append({"username": u, "ipaddress": ip, "date": _iso_date(f.get("date") or f.get("data") or "")})
        return out

    # Kolumnowy eksport maklerski: znajdź wiersz nagłówka z kolumną IP.
    ip_idx = date_idx = None
    for row in grid:
        found_ip = found_date = None
        for j, c in enumerate(row):
            if c is None:
                continue
            nm = str(c).strip().lower()
            if found_ip is None and nm in _IP_COLS:
                found_ip = j
            if found_date is None and nm in _DATE_COLS:
                found_date = j
        if found_ip is not None:
            ip_idx, date_idx = found_ip, found_date
            break
    if ip_idx is None:
        return []
    hdr_seen, out = False, []
    for row in grid:
        if not hdr_seen:
            if ip_idx < len(row) and row[ip_idx] is not None and str(row[ip_idx]).strip().lower() in _IP_COLS:
                hdr_seen = True
            continue
        if ip_idx >= len(row) or row[ip_idx] is None:
            continue
        ip = str(row[ip_idx]).strip()
        if not _IPV4.match(ip):
            continue
        d = str(row[date_idx]).strip() if (date_idx is not None and date_idx < len(row) and row[date_idx] is not None) else ""
        out.append({"username": ent, "ipaddress": ip, "date": _iso_date(d)})
    return out


def _iso_date(raw: str) -> str | None:
    """Normalizuje datę logowania do ISO YYYY-MM-DD (tolerancyjnie: ISO, DD.MM.YYYY,
    DD-MM-YYYY, DD/MM/YYYY, 'YYYY-MM-DD hh:mm:ss'). Nieczytelna → None."""
    s = (raw or "").strip().split(" ")[0].split("T")[0]
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return s
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{4})$", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def ip_correlation(rows: list[dict], max_users_per_ip: int = 8) -> dict:
    """Pary użytkowników dzielących adresy IP + zdarzenia logowań ze wspólnych IP.

    Bierzemy tylko adresy współdzielone przez 2..`max_users_per_ip` użytkowników —
    IP użyty przez wielu (proxy/publiczny) nie jest znamienny. Zwraca pary z liczbą
    wspólnych adresów, statystyki zbiorcze oraz `events` — unikalne (data, IP,
    użytkownik) WYŁĄCZNIE dla wspólnych adresów (materiał wykresu „data × IP"
    w formie jak wykres nr 6 analizy specjalisty: nałożenie symboli = wspólne IP).
    """
    ip_users: dict[str, set] = defaultdict(set)
    ip_user_dates: dict[tuple, set] = defaultdict(set)
    for r in rows:
        f = _fields(r)
        u = f.get("username") or f.get("user") or f.get("login")
        ip = f.get("ipaddress") or f.get("ip") or f.get("ipaddr") or f.get("adres ip")
        if not (u and ip):
            continue
        ip_users[ip].add(u)
        d = _iso_date(f.get("date") or f.get("data") or "")
        if d:
            ip_user_dates[(ip, u)].add(d)

    pairs: dict[tuple, set] = defaultdict(set)
    shared_ips: list[str] = []
    for ip, users in ip_users.items():
        if not (2 <= len(users) <= max_users_per_ip):
            continue
        shared_ips.append(ip)
        us = sorted(users)
        for i in range(len(us)):
            for j in range(i + 1, len(us)):
                pairs[(us[i], us[j])].add(ip)

    out = [
        {"user_a": a, "user_b": b, "n_shared": len(ips), "shared_ips": sorted(ips)}
        for (a, b), ips in pairs.items()
    ]
    out.sort(key=lambda x: (-x["n_shared"], x["user_a"], x["user_b"]))

    shared_set = set(shared_ips)
    events = [
        {"date": d, "ip": ip, "user": u}
        for (ip, u), dates in ip_user_dates.items()
        if ip in shared_set
        for d in dates
    ]
    events.sort(key=lambda e: (e["date"], e["ip"], e["user"]))
    return {
        "pairs": out,
        "events": events,
        "shared_ip_count": len(shared_ips),
        "ip_count": len(ip_users),
        "user_count": len({u for us in ip_users.values() for u in us}),
    }
